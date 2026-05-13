#!/usr/bin/env python3
"""
Экспорт звонков в Excel в формате карточек.
Каждый звонок — вертикальный блок строк, транскрипт читается сверху вниз.
"""

import json
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DATA_DIR = Path(__file__).parent.parent / "data"
CALLS_DIR = DATA_DIR / "calls"
TRANSCRIPTS_DIR = DATA_DIR / "transcripts"
ANALYSIS_DIR = DATA_DIR / "analysis"
REPORTS_DIR = DATA_DIR / "reports"

REPORTS_DIR.mkdir(parents=True, exist_ok=True)

# (фон, цвет текста) для каждого score
SCORE_COLORS = {
    1: ("FF4444", "FFFFFF"),
    2: ("FF8C00", "FFFFFF"),
    3: ("FFD700", "333333"),
    4: ("5DBB5D", "FFFFFF"),
    5: ("2E8B57", "FFFFFF"),
}
COLOR_UNKNOWN    = ("AAAAAA", "FFFFFF")
COLOR_AGENT_BG   = ("2C3E6B", "FFFFFF")
COLOR_TITLE_BG   = "2E4057"
COLOR_STATS_BG   = "D9E8F5"
COLOR_INFO_BG    = "F5F8FF"
COLOR_AI_BG      = "EEF6EE"
COLOR_ISSUES_BG  = "FFF3E0"
COLOR_TR_HEAD_BG = "E8E8E8"
COLOR_TR_TEXT_BG = "FAFAFA"
COLOR_SEPARATOR  = "CCCCCC"

LONG_CALL_SEC      = 180   # > 3 мин
VERY_LONG_CALL_SEC = 300   # > 5 мин
PRODUCTIVE_OUTCOMES = {"booked", "interested"}

OUTCOME_RU = {
    "booked":            "Бронь ✅",
    "interested":        "Интерес",
    "callback":          "Перезвон",
    "info_only":         "Информация",
    "not_interested":    "Отказ",
    "complaint":         "Жалоба",
    "wrong_number":      "Ошибка номера",
    "client_unavailable":"Клиент занят",
    "unknown":           "—",
    "no_transcript":     "Нет транскрипта",
}


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=10, color="000000", name="Calibri"):
    return Font(name=name, size=size, bold=bold, color=color)


def _align(wrap=True, valign="top", halign="left", indent=1):
    return Alignment(wrap_text=wrap, vertical=valign, horizontal=halign, indent=indent)


def build_transcript_text(transcript_data: dict) -> str:
    utterances = transcript_data.get("utterances", [])
    if utterances:
        lines = []
        for u in utterances:
            secs = int(u.get("start", 0))
            m, s = divmod(secs, 60)
            speaker = "Менеджер" if u.get("speaker", 0) == 0 else "Клиент  "
            lines.append(f"[{m}:{s:02d}]  {speaker}: {u.get('text', '')}")
        return "\n".join(lines)
    return transcript_data.get("transcript", "")


def load_phones_map() -> dict:
    f = DATA_DIR / "managers.json"
    return json.loads(f.read_text()) if f.exists() else {}


def load_agents() -> dict:
    f = DATA_DIR / "agents.json"
    return json.loads(f.read_text()) if f.exists() else {}


def export_problem_calls_excel(
    date: datetime,
    phones_map: dict,
    min_score_threshold: int = 2,
    out_path: Path = None,
) -> Path | None:
    calls_file = CALLS_DIR / f"{date.strftime('%Y-%m-%d')}.json"
    if not calls_file.exists():
        return None

    calls = json.loads(calls_file.read_text())
    agents = load_agents()
    rows = []

    for call in calls:
        pid = call.get("public_id")
        af = ANALYSIS_DIR / f"{pid}.json"
        tf = TRANSCRIPTS_DIR / f"{pid}.json"

        analysis       = json.loads(af.read_text()) if af.exists() else {}
        transcript_data = json.loads(tf.read_text()) if tf.exists() else {}

        call_type = call.get("call_type", "")
        if call_type == "outbound":
            manager_num = call.get("caller_number", "")
            client_num  = call.get("target_number", "")
        else:
            manager_num = call.get("answered_phone_number") or call.get("caller_number", "")
            client_num  = call.get("caller_number", "")

        manager_name = phones_map.get(manager_num, manager_num or "—")

        start_raw = call.get("start_time") or call.get("registration_time") or ""
        try:
            start_dt = datetime.fromisoformat(start_raw.replace("Z", "+00:00"))
            time_str = start_dt.strftime("%H:%M")
        except Exception:
            time_str = start_raw[:16] if start_raw else "—"

        dur  = call.get("duration_sec", 0) or 0
        wait = call.get("wait_duration_sec", 0) or 0
        dur_str = f"{dur // 60}:{dur % 60:02d}"

        score       = analysis.get("quality_score")
        raw_outcome = analysis.get("outcome", "")
        outcome     = OUTCOME_RU.get(raw_outcome, "—")
        direction   = analysis.get("direction") or "—"
        issues      = ", ".join(analysis.get("issues", []))
        highlights  = analysis.get("manager_highlights", "") or ""
        transcript_text = build_transcript_text(transcript_data) if transcript_data else "Нет записи"

        sip = call.get("sip_status", "")
        status_ru = {"answer": "Отвечен", "no_answer": "Не отвечен",
                     "cancel": "Отменён",  "busy": "Занято"}.get(sip, sip)

        client_unavailable = raw_outcome in ("client_unavailable", "wrong_number")
        long_no_result = (
            dur >= LONG_CALL_SEC
            and raw_outcome not in PRODUCTIVE_OUTCOMES
            and not client_unavailable
        )
        is_agent = bool(analysis.get("is_agent")) or client_num in agents

        rows.append({
            "time": time_str, "manager": manager_name,
            "call_type": "Вход." if call_type == "inbound" else "Исход.",
            "status": status_ru, "client": client_num,
            "duration": dur_str, "wait": f"{wait}с",
            "score": score, "outcome": outcome,
            "direction": direction, "issues": issues,
            "highlights": highlights, "transcript": transcript_text,
            "_score_num": score or 99,
            "_client_unavailable": client_unavailable,
            "_dur_sec": dur,
            "_long_no_result": long_no_result,
            "_is_agent": is_agent,
        })

    if not rows:
        return None

    # Сортировка: проблемные → длинные без результата → остальные; внутри — по времени
    def _sort_key(r):
        if r["_score_num"] <= min_score_threshold and not r["_client_unavailable"]:
            g = 0
        elif r["_long_no_result"]:
            g = 1
        else:
            g = 2
        return (g, r["time"])

    rows.sort(key=_sort_key)

    total          = len(rows)
    answered       = sum(1 for r in rows if r["status"] == "Отвечен")
    problematic    = sum(1 for r in rows if r["_score_num"] <= min_score_threshold and not r["_client_unavailable"])
    long_no_result = sum(1 for r in rows if r["_long_no_result"])

    # ── Excel ────────────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = f"Звонки {date.strftime('%d.%m.%Y')}"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 145

    row = 1

    # Заголовок
    c = ws.cell(row=row, column=1,
                value=f"📞 Аналитика звонков GoTrips — {date.strftime('%d.%m.%Y')}")
    c.font      = _font(bold=True, size=14, color="FFFFFF")
    c.fill      = _fill(COLOR_TITLE_BG)
    c.alignment = _align(wrap=False, valign="center", halign="center", indent=0)
    ws.row_dimensions[row].height = 34
    row += 1

    # Статистика
    c = ws.cell(row=row, column=1, value=(
        f"Всего: {total}  │  Отвечено: {answered}  │  "
        f"Проблемных (score ≤{min_score_threshold}): {problematic}  │  "
        f"⚡ Длинных без результата (>3 мин): {long_no_result}"
    ))
    c.font      = _font(size=10, color="444444")
    c.fill      = _fill(COLOR_STATS_BG)
    c.alignment = _align(wrap=False, valign="center", halign="center", indent=0)
    ws.row_dimensions[row].height = 22
    row += 1

    # ── Критерии оценки ──────────────────────────────────────────────────────
    criteria_lines = [
        "  КРИТЕРИИ ОЦЕНКИ AI (качество работы менеджера):",
        "  ★★★★★  5 — Отлично:  выявил потребности, обработал возражения, договорился о следующем шаге",
        "  ★★★★☆  4 — Хорошо:   дал чёткий ответ / профессионально провёл B2B-звонок / правильно отработал нестандартную ситуацию",
        "  ★★★☆☆  3 — Норма:    ответил на вопросы, но не развил диалог и не продвинул к покупке",
        "  ★★☆☆☆  2 — Слабо:    формальный ответ, без вовлечения, упустил явную возможность",
        "  ★☆☆☆☆  1 — Критично: грубость, игнор вопросов, бросил трубку, конфликт, потеря клиента",
        "  ─────────────────────────────────────────────────────────────────────────────────────",
        "  ИСХОДЫ:  Бронь ✅ — забронировал тур   │  Интерес — клиент заинтересован, продолжает общение",
        "           Перезвон — договорились созвониться позже   │  Информация — ответили на вопрос, без продажи",
        "           Отказ — клиент отказался   │  Клиент занят — не мог говорить (не вина менеджера)",
        "  ─────────────────────────────────────────────────────────────────────────────────────",
        "  ⚡ — звонок длился более 3 минут, но не завершился бронью или явным интересом к туру",
        "  🔴 — звонок длился более 5 минут без результата (приоритет для разбора с менеджером)",
    ]

    COLOR_CRITERIA_BG    = "F9F9F9"
    COLOR_CRITERIA_LABEL = "EEEEEE"

    for i, line in enumerate(criteria_lines):
        c = ws.cell(row=row, column=1, value=line)
        if i == 0:
            c.font = _font(size=9, bold=True, color="2E4057")
            c.fill = _fill(COLOR_CRITERIA_LABEL)
            ws.row_dimensions[row].height = 16
        elif "─────" in line:
            c.font = _font(size=8, color="AAAAAA")
            c.fill = _fill(COLOR_CRITERIA_BG)
            ws.row_dimensions[row].height = 10
        else:
            c.font = _font(size=9, color="333333")
            c.fill = _fill(COLOR_CRITERIA_BG)
            ws.row_dimensions[row].height = 15
        c.alignment = _align(wrap=False, valign="center", halign="left", indent=0)
        row += 1

    row += 1  # отступ перед карточками

    # Закрепляем строки 1 и 2 (заголовок + статистика)
    ws.freeze_panes = "A3"

    # ── Карточки ─────────────────────────────────────────────────────────────
    for call_num, r in enumerate(rows, start=1):
        score = r["score"]
        if r.get("_is_agent"):
            bg_color, fg_color = COLOR_AGENT_BG
        else:
            bg_color, fg_color = SCORE_COLORS.get(score, COLOR_UNKNOWN) if score else COLOR_UNKNOWN

        # Звёзды оценки
        stars = ("★" * score + "☆" * (5 - score)) if score else "—"

        # Маркер длинного звонка
        if r["_dur_sec"] >= VERY_LONG_CALL_SEC and r["_long_no_result"]:
            long_mark = "   ⚡🔴 ДЛИННЫЙ БЕЗ РЕЗУЛЬТАТА"
        elif r["_long_no_result"]:
            long_mark = "   ⚡ Длинный без результата"
        else:
            long_mark = ""

        agent_mark = "   🤝 АГЕНТ" if r.get("_is_agent") else ""

        # ── Строка 1: шапка карточки (цвет по оценке) ────────────────────
        header = (
            f"  #{call_num}   {r['time']}   {r['manager']}   {r['call_type']}   "
            f"⏱ {r['duration']}   {stars}   {r['outcome']}   {r['direction']}"
            f"{long_mark}{agent_mark}"
        )
        c = ws.cell(row=row, column=1, value=header)
        c.font      = _font(bold=True, size=12, color=fg_color)
        c.fill      = _fill(bg_color)
        c.alignment = _align(wrap=False, valign="center", halign="left", indent=1)
        ws.row_dimensions[row].height = 28
        row += 1

        # ── Строка 2: клиент / ожидание / статус ─────────────────────────
        c = ws.cell(row=row, column=1,
                    value=f"  Клиент: {r['client']}   │   Ожидание: {r['wait']}   │   {r['status']}")
        c.font      = _font(size=10, color="555555")
        c.fill      = _fill(COLOR_INFO_BG)
        c.alignment = _align(wrap=False, valign="center", halign="left", indent=1)
        ws.row_dimensions[row].height = 20
        row += 1

        # ── Строка 3: оценка менеджера от AI ─────────────────────────────
        if highlights:
            c = ws.cell(row=row, column=1, value=f"  🤖  {highlights}")
            c.font      = _font(size=11, color="1A5276")
            c.fill      = _fill(COLOR_AI_BG)
            c.alignment = _align(wrap=True, valign="top", halign="left", indent=1)
            h = max(int(len(highlights) / 125 + 1) * 16, 22)
            ws.row_dimensions[row].height = h
            row += 1

        # ── Строка 4: проблемы (если есть) ───────────────────────────────
        if r["issues"]:
            c = ws.cell(row=row, column=1, value=f"  ⚠️   {r['issues']}")
            c.font      = _font(size=9, color="8B4513", bold=True)
            c.fill      = _fill(COLOR_ISSUES_BG)
            c.alignment = _align(wrap=True, valign="top", halign="left", indent=1)
            ws.row_dimensions[row].height = 18
            row += 1

        # ── Строка 5: заголовок транскрипта ──────────────────────────────
        c = ws.cell(row=row, column=1, value="  📝  ТРАНСКРИПТ")
        c.font      = _font(size=9, bold=True, color="444444")
        c.fill      = _fill(COLOR_TR_HEAD_BG)
        c.alignment = _align(wrap=False, valign="center", halign="left", indent=1)
        ws.row_dimensions[row].height = 16
        row += 1

        # ── Строка 6+: текст транскрипта ──────────────────────────────────
        transcript = r["transcript"]
        c = ws.cell(row=row, column=1, value=transcript)
        c.font      = Font(name="Courier New", size=9, color="222222")
        c.fill      = _fill(COLOR_TR_TEXT_BG)
        c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left", indent=2)
        lines = transcript.count("\n") + 1
        ws.row_dimensions[row].height = min(lines * 14, 500)
        row += 1

        # ── Разделитель ───────────────────────────────────────────────────
        c = ws.cell(row=row, column=1, value="")
        c.border = Border(bottom=Side(style="medium", color=COLOR_SEPARATOR))
        ws.row_dimensions[row].height = 10
        row += 1

    if out_path is None:
        out_path = REPORTS_DIR / f"calls_{date.strftime('%Y-%m-%d')}.xlsx"

    wb.save(str(out_path))
    return out_path


def _build_rows_for_calls(calls: list, phones_map: dict, agents: dict, date_prefix: str = "") -> list:
    """Build row dicts for a list of CDR calls. date_prefix added to time field (e.g. '01.04 ')."""
    rows = []
    for call in calls:
        pid = call.get("public_id")
        af = ANALYSIS_DIR / f"{pid}.json"
        tf = TRANSCRIPTS_DIR / f"{pid}.json"

        analysis        = json.loads(af.read_text()) if af.exists() else {}
        transcript_data = json.loads(tf.read_text()) if tf.exists() else {}

        call_type = call.get("call_type", "")
        if call_type == "outbound":
            manager_num = call.get("caller_number", "")
            client_num  = call.get("target_number", "")
        else:
            manager_num = call.get("answered_phone_number") or call.get("caller_number", "")
            client_num  = call.get("caller_number", "")

        manager_name = phones_map.get(manager_num, manager_num or "—")

        start_raw = call.get("start_time") or call.get("registration_time") or ""
        try:
            start_dt = datetime.fromisoformat(start_raw.replace("Z", "+00:00"))
            time_str = date_prefix + start_dt.strftime("%H:%M")
        except Exception:
            time_str = date_prefix + (start_raw[:5] if start_raw else "—")

        dur  = call.get("duration_sec", 0) or 0
        wait = call.get("wait_duration_sec", 0) or 0
        dur_str = f"{dur // 60}:{dur % 60:02d}"

        score       = analysis.get("quality_score")
        raw_outcome = analysis.get("outcome", "")
        outcome     = OUTCOME_RU.get(raw_outcome, "—")
        direction   = analysis.get("direction") or "—"
        issues      = ", ".join(analysis.get("issues", []))
        highlights  = analysis.get("manager_highlights", "") or ""
        transcript_text = build_transcript_text(transcript_data) if transcript_data else "Нет записи"

        sip = call.get("sip_status", "")
        status_ru = {"answer": "Отвечен", "no_answer": "Не отвечен",
                     "cancel": "Отменён", "busy": "Занято"}.get(sip, sip)

        client_unavailable = raw_outcome in ("client_unavailable", "wrong_number")
        long_no_result = (
            dur >= LONG_CALL_SEC
            and raw_outcome not in PRODUCTIVE_OUTCOMES
            and not client_unavailable
        )
        is_agent = bool(analysis.get("is_agent")) or client_num in agents

        rows.append({
            "time": time_str, "manager": manager_name,
            "call_type": "Вход." if call_type == "inbound" else "Исход.",
            "status": status_ru, "client": client_num,
            "duration": dur_str, "wait": f"{wait}с",
            "score": score, "outcome": outcome,
            "direction": direction, "issues": issues,
            "highlights": highlights, "transcript": transcript_text,
            "_score_num": score or 99,
            "_client_unavailable": client_unavailable,
            "_dur_sec": dur,
            "_long_no_result": long_no_result,
            "_is_agent": is_agent,
            "_sort_time": start_raw,
        })
    return rows


def _write_rows_to_sheet(ws, rows: list, min_score_threshold: int = 2, row_start: int = 1) -> int:
    """Write card rows to worksheet starting at row_start. Returns next row number."""
    row = row_start
    for call_num, r in enumerate(rows, start=1):
        score = r["score"]
        if r.get("_is_agent"):
            bg_color, fg_color = COLOR_AGENT_BG
        else:
            bg_color, fg_color = SCORE_COLORS.get(score, COLOR_UNKNOWN) if score else COLOR_UNKNOWN

        stars = ("★" * score + "☆" * (5 - score)) if score else "—"

        if r["_dur_sec"] >= VERY_LONG_CALL_SEC and r["_long_no_result"]:
            long_mark = "   ⚡🔴 ДЛИННЫЙ БЕЗ РЕЗУЛЬТАТА"
        elif r["_long_no_result"]:
            long_mark = "   ⚡ Длинный без результата"
        else:
            long_mark = ""

        agent_mark = "   🤝 АГЕНТ" if r.get("_is_agent") else ""

        header = (
            f"  #{call_num}   {r['time']}   {r['manager']}   {r['call_type']}   "
            f"⏱ {r['duration']}   {stars}   {r['outcome']}   {r['direction']}"
            f"{long_mark}{agent_mark}"
        )
        c = ws.cell(row=row, column=1, value=header)
        c.font      = _font(bold=True, size=12, color=fg_color)
        c.fill      = _fill(bg_color)
        c.alignment = _align(wrap=False, valign="center", halign="left", indent=1)
        ws.row_dimensions[row].height = 28
        row += 1

        c = ws.cell(row=row, column=1,
                    value=f"  Клиент: {r['client']}   │   Ожидание: {r['wait']}   │   {r['status']}")
        c.font      = _font(size=10, color="555555")
        c.fill      = _fill(COLOR_INFO_BG)
        c.alignment = _align(wrap=False, valign="center", halign="left", indent=1)
        ws.row_dimensions[row].height = 20
        row += 1

        if r["highlights"]:
            c = ws.cell(row=row, column=1, value=f"  🤖  {r['highlights']}")
            c.font      = _font(size=11, color="1A5276")
            c.fill      = _fill(COLOR_AI_BG)
            c.alignment = _align(wrap=True, valign="top", halign="left", indent=1)
            h = max(int(len(r["highlights"]) / 125 + 1) * 16, 22)
            ws.row_dimensions[row].height = h
            row += 1

        if r["issues"]:
            c = ws.cell(row=row, column=1, value=f"  ⚠️   {r['issues']}")
            c.font      = _font(size=9, color="8B4513", bold=True)
            c.fill      = _fill(COLOR_ISSUES_BG)
            c.alignment = _align(wrap=True, valign="top", halign="left", indent=1)
            ws.row_dimensions[row].height = 18
            row += 1

        c = ws.cell(row=row, column=1, value="  📝  ТРАНСКРИПТ")
        c.font      = _font(size=9, bold=True, color="444444")
        c.fill      = _fill(COLOR_TR_HEAD_BG)
        c.alignment = _align(wrap=False, valign="center", halign="left", indent=1)
        ws.row_dimensions[row].height = 16
        row += 1

        transcript = r["transcript"]
        c = ws.cell(row=row, column=1, value=transcript)
        c.font      = Font(name="Courier New", size=9, color="222222")
        c.fill      = _fill(COLOR_TR_TEXT_BG)
        c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left", indent=2)
        lines = transcript.count("\n") + 1
        ws.row_dimensions[row].height = min(lines * 14, 500)
        row += 1

        c = ws.cell(row=row, column=1, value="")
        c.border = Border(bottom=Side(style="medium", color=COLOR_SEPARATOR))
        ws.row_dimensions[row].height = 10
        row += 1

    return row


def export_period_excel(
    dates: list,
    phones_map: dict,
    period_label: str,
    out_path: Path = None,
) -> Path | None:
    """Export all calls for a list of dates into one Excel workbook (monthly report)."""
    agents = load_agents()
    all_rows = []

    for d in dates:
        calls_file = CALLS_DIR / f"{d.strftime('%Y-%m-%d')}.json"
        if not calls_file.exists():
            continue
        calls = json.loads(calls_file.read_text())
        date_prefix = d.strftime("%d.%m ")
        all_rows.extend(_build_rows_for_calls(calls, phones_map, agents, date_prefix))

    if not all_rows:
        return None

    min_score_threshold = 2

    def _sort_key(r):
        if r["_score_num"] <= min_score_threshold and not r["_client_unavailable"]:
            g = 0
        elif r["_long_no_result"]:
            g = 1
        else:
            g = 2
        return (g, r.get("_sort_time", ""))

    all_rows.sort(key=_sort_key)

    total          = len(all_rows)
    answered       = sum(1 for r in all_rows if r["status"] == "Отвечен")
    problematic    = sum(1 for r in all_rows if r["_score_num"] <= min_score_threshold and not r["_client_unavailable"])
    long_no_result = sum(1 for r in all_rows if r["_long_no_result"])

    wb = Workbook()
    ws = wb.active
    ws.title = f"Звонки {period_label}"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 145

    row = 1

    c = ws.cell(row=row, column=1,
                value=f"📞 Аналитика звонков GoTrips — {period_label}")
    c.font      = _font(bold=True, size=14, color="FFFFFF")
    c.fill      = _fill(COLOR_TITLE_BG)
    c.alignment = _align(wrap=False, valign="center", halign="center", indent=0)
    ws.row_dimensions[row].height = 34
    row += 1

    c = ws.cell(row=row, column=1, value=(
        f"Всего: {total}  │  Отвечено: {answered}  │  "
        f"Проблемных (score ≤{min_score_threshold}): {problematic}  │  "
        f"Длинных без результата (>3 мин): {long_no_result}  │  "
        f"Дней с данными: {len(dates)}"
    ))
    c.font      = _font(size=10, color="444444")
    c.fill      = _fill(COLOR_STATS_BG)
    c.alignment = _align(wrap=False, valign="center", halign="center", indent=0)
    ws.row_dimensions[row].height = 22
    row += 1

    ws.freeze_panes = "A3"
    row += 1

    _write_rows_to_sheet(ws, all_rows, min_score_threshold, row_start=row)

    if out_path is None:
        out_path = REPORTS_DIR / f"calls_{period_label.replace(' ', '_')}.xlsx"

    wb.save(str(out_path))
    return out_path


def main():
    import argparse
    from datetime import timedelta

    parser = argparse.ArgumentParser()
    parser.add_argument("--date", help="YYYY-MM-DD (default: yesterday)")
    args = parser.parse_args()

    target_date = (
        datetime.strptime(args.date, "%Y-%m-%d") if args.date
        else datetime.now() - timedelta(days=1)
    )

    phones = load_phones_map()
    out = export_problem_calls_excel(target_date, phones)
    if out:
        print(f"[export] Saved: {out}")
    else:
        print("[export] No data")


if __name__ == "__main__":
    main()
